"""ExportController for data export functionality."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import TYPE_CHECKING, Any, ClassVar, Literal

from litestar import Controller, get, post
from litestar.exceptions import NotFoundException, ValidationException
from litestar.response import Stream
from litestar.status_codes import HTTP_200_OK

# Litestar requires these imports at runtime for dependency injection type resolution
from sqlalchemy import inspect, select
from sqlalchemy.ext.asyncio import AsyncSession  # noqa: TC002

from litestar_admin.registry import ModelRegistry  # noqa: TC001

if TYPE_CHECKING:
    from collections.abc import AsyncGenerator

__all__ = [
    "BulkExportRequest",
    "ExportController",
]

# Supported export formats
ExportFormat = Literal["csv", "json", "xlsx"]
SUPPORTED_FORMATS: frozenset[str] = frozenset({"csv", "json", "xlsx"})

# Content types for export formats
CONTENT_TYPES: dict[str, str] = {
    "csv": "text/csv",
    "json": "application/json",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

# Streaming configuration
CHUNK_SIZE = 1000  # Number of records to process per chunk


@dataclass
class BulkExportRequest:
    """Request body for bulk export operation.

    Attributes:
        ids: List of record IDs to export.
        format: Export format (csv, json, or xlsx).
    """

    ids: list[Any]
    format: ExportFormat = "csv"


def _get_openpyxl() -> Any:
    """Import and return openpyxl module.

    This function lazily imports openpyxl only when XLSX export is requested,
    allowing the library to remain an optional dependency.

    Returns:
        The openpyxl module.

    Raises:
        ValidationException: If openpyxl is not installed.
    """
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        msg = "XLSX export requires the 'openpyxl' package. Install it with: pip install litestar-admin[excel]"
        raise ValidationException(msg) from None


class ExportController(Controller):
    """Controller for exporting model data.

    Provides endpoints for exporting records in CSV, JSON, or XLSX format.
    Supports both full exports and selective bulk exports.

    Note:
        XLSX export requires the optional 'excel' dependency (openpyxl).
        Install with: pip install litestar-admin[excel]

    Example:
        The controller is automatically registered by AdminPlugin.
        Access endpoints at:
        - GET /admin/api/models/{model}/export?format=csv
        - GET /admin/api/models/{model}/export?format=xlsx
        - POST /admin/api/models/{model}/bulk/export
    """

    path = "/api/models"
    tags: ClassVar[list[str]] = ["Export"]

    @get(
        "/{model_name:str}/export",
        status_code=HTTP_200_OK,
        summary="Export all records",
        description="Export all records from a model in CSV, JSON, or XLSX format.",
    )
    async def export_all(
        self,
        model_name: str,
        admin_registry: ModelRegistry,
        db_session: AsyncSession,
        format: ExportFormat = "csv",  # noqa: A002
    ) -> Stream:
        """Export all records from a model.

        Args:
            model_name: The name of the model to export.
            admin_registry: The model registry containing all registered views.
            db_session: The database session.
            format: Export format (csv, json, or xlsx). Defaults to csv.

        Returns:
            Streaming response with exported data.

        Raises:
            NotFoundException: If the model is not found.
            ValidationException: If the format is not supported.
        """
        self._validate_format(format)
        view_class = self._get_view_class(admin_registry, model_name)

        if not view_class.can_export:
            msg = f"Export is not allowed for model '{model_name}'"
            raise ValidationException(msg)

        model = view_class.model
        columns = view_class.get_list_columns()

        async def generate_export() -> AsyncGenerator[bytes, None]:
            """Generate export data in chunks."""
            if format == "csv":
                async for chunk in self._stream_csv(db_session, model, columns):
                    yield chunk
            elif format == "xlsx":
                async for chunk in self._stream_xlsx(db_session, model, columns, model_name):
                    yield chunk
            else:
                async for chunk in self._stream_json(db_session, model, columns):
                    yield chunk

        content_type = CONTENT_TYPES.get(format, "application/octet-stream")
        filename = f"{model_name}_export.{format}"

        return Stream(
            generate_export(),
            media_type=content_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    @post(
        "/{model_name:str}/bulk/export",
        status_code=HTTP_200_OK,
        summary="Export selected records",
        description="Export selected records by their IDs in CSV, JSON, or XLSX format.",
    )
    async def export_selected(
        self,
        model_name: str,
        data: BulkExportRequest,
        admin_registry: ModelRegistry,
        db_session: AsyncSession,
    ) -> Stream:
        """Export selected records by their IDs.

        Args:
            model_name: The name of the model to export.
            data: Request containing IDs and format.
            admin_registry: The model registry containing all registered views.
            db_session: The database session.

        Returns:
            Streaming response with exported data.

        Raises:
            NotFoundException: If the model is not found.
            ValidationException: If the format is not supported or no IDs provided.
        """
        self._validate_format(data.format)
        view_class = self._get_view_class(admin_registry, model_name)

        if not view_class.can_export:
            msg = f"Export is not allowed for model '{model_name}'"
            raise ValidationException(msg)

        if not data.ids:
            msg = "No IDs provided for export"
            raise ValidationException(msg)

        model = view_class.model
        columns = view_class.get_list_columns()

        # Get primary key column
        pk_column = self._get_primary_key_column(model)

        async def generate_export() -> AsyncGenerator[bytes, None]:
            """Generate export data for selected records."""
            if data.format == "csv":
                async for chunk in self._stream_csv_by_ids(db_session, model, columns, pk_column, data.ids):
                    yield chunk
            elif data.format == "xlsx":
                async for chunk in self._stream_xlsx_by_ids(
                    db_session, model, columns, pk_column, data.ids, model_name
                ):
                    yield chunk
            else:
                async for chunk in self._stream_json_by_ids(db_session, model, columns, pk_column, data.ids):
                    yield chunk

        content_type = CONTENT_TYPES.get(data.format, "application/octet-stream")
        filename = f"{model_name}_export.{data.format}"

        return Stream(
            generate_export(),
            media_type=content_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    @staticmethod
    def _validate_format(export_format: str) -> None:
        """Validate the export format.

        Args:
            export_format: The requested export format.

        Raises:
            ValidationException: If the format is not supported.
        """
        if export_format not in SUPPORTED_FORMATS:
            msg = f"Unsupported export format: '{export_format}'. Supported formats: {', '.join(SUPPORTED_FORMATS)}"
            raise ValidationException(msg)

    @staticmethod
    def _get_view_class(
        registry: ModelRegistry,
        model_name: str,
    ) -> Any:
        """Get the view class for a model by name.

        Args:
            registry: The model registry.
            model_name: The model name to look up.

        Returns:
            The view class for the model.

        Raises:
            NotFoundException: If the model is not found.
        """
        if not registry.has_model_by_name(model_name):
            msg = f"Model '{model_name}' not found"
            raise NotFoundException(msg)
        return registry.get_view_by_name(model_name)

    @staticmethod
    def _get_primary_key_column(model: type[Any]) -> str:
        """Get the primary key column name for a model.

        Args:
            model: The SQLAlchemy model class.

        Returns:
            The primary key column name.

        Raises:
            ValidationException: If no primary key is found.
        """
        mapper = inspect(model)
        pk_columns = mapper.primary_key
        if not pk_columns:
            msg = f"Model {model.__name__} has no primary key"
            raise ValidationException(msg)
        return pk_columns[0].name

    @staticmethod
    def _record_to_dict(record: Any, columns: list[str]) -> dict[str, Any]:
        """Convert a record to a dictionary with specified columns.

        Args:
            record: The SQLAlchemy model instance.
            columns: List of column names to include.

        Returns:
            Dictionary with column values.
        """
        result: dict[str, Any] = {}
        for column in columns:
            value = getattr(record, column, None)
            # Handle common non-serializable types
            if hasattr(value, "isoformat"):
                # datetime, date, time objects
                value = value.isoformat()
            elif hasattr(value, "__dict__") and not isinstance(value, (str, bytes, int, float, bool, type(None))):
                # Related objects - just use string representation
                value = str(value)
            result[column] = value
        return result

    async def _stream_csv(
        self,
        session: AsyncSession,
        model: type[Any],
        columns: list[str],
    ) -> AsyncGenerator[bytes, None]:
        """Stream CSV data for all records.

        Args:
            session: The database session.
            model: The SQLAlchemy model class.
            columns: List of column names to export.

        Yields:
            CSV data in chunks.
        """
        # Write header
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        yield output.getvalue().encode("utf-8")

        # Stream records in chunks
        offset = 0
        while True:
            query = select(model).offset(offset).limit(CHUNK_SIZE)
            result = await session.scalars(query)
            records = result.all()

            if not records:
                break

            output = StringIO()
            writer = csv.writer(output)
            for record in records:
                row_data = self._record_to_dict(record, columns)
                writer.writerow([row_data.get(col) for col in columns])
            yield output.getvalue().encode("utf-8")

            offset += CHUNK_SIZE

            # If we got fewer records than the chunk size, we're done
            if len(records) < CHUNK_SIZE:
                break

    async def _stream_json(
        self,
        session: AsyncSession,
        model: type[Any],
        columns: list[str],
    ) -> AsyncGenerator[bytes, None]:
        """Stream JSON data for all records.

        Args:
            session: The database session.
            model: The SQLAlchemy model class.
            columns: List of column names to export.

        Yields:
            JSON data in chunks (as a JSON array).
        """
        yield b"["

        offset = 0
        first_chunk = True
        while True:
            query = select(model).offset(offset).limit(CHUNK_SIZE)
            result = await session.scalars(query)
            records = result.all()

            if not records:
                break

            for record in records:
                row_data = self._record_to_dict(record, columns)
                prefix = "" if first_chunk else ","
                first_chunk = False
                yield (prefix + json.dumps(row_data, default=str)).encode("utf-8")

            offset += CHUNK_SIZE

            if len(records) < CHUNK_SIZE:
                break

        yield b"]"

    async def _stream_csv_by_ids(
        self,
        session: AsyncSession,
        model: type[Any],
        columns: list[str],
        pk_column: str,
        ids: list[Any],
    ) -> AsyncGenerator[bytes, None]:
        """Stream CSV data for selected records.

        Args:
            session: The database session.
            model: The SQLAlchemy model class.
            columns: List of column names to export.
            pk_column: The primary key column name.
            ids: List of record IDs to export.

        Yields:
            CSV data in chunks.
        """
        # Write header
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        yield output.getvalue().encode("utf-8")

        # Process IDs in chunks
        for i in range(0, len(ids), CHUNK_SIZE):
            chunk_ids = ids[i : i + CHUNK_SIZE]
            pk_attr = getattr(model, pk_column)
            query = select(model).where(pk_attr.in_(chunk_ids))
            result = await session.scalars(query)
            records = result.all()

            if records:
                output = StringIO()
                writer = csv.writer(output)
                for record in records:
                    row_data = self._record_to_dict(record, columns)
                    writer.writerow([row_data.get(col) for col in columns])
                yield output.getvalue().encode("utf-8")

    async def _stream_json_by_ids(
        self,
        session: AsyncSession,
        model: type[Any],
        columns: list[str],
        pk_column: str,
        ids: list[Any],
    ) -> AsyncGenerator[bytes, None]:
        """Stream JSON data for selected records.

        Args:
            session: The database session.
            model: The SQLAlchemy model class.
            columns: List of column names to export.
            pk_column: The primary key column name.
            ids: List of record IDs to export.

        Yields:
            JSON data in chunks (as a JSON array).
        """
        yield b"["

        first_record = True
        for i in range(0, len(ids), CHUNK_SIZE):
            chunk_ids = ids[i : i + CHUNK_SIZE]
            pk_attr = getattr(model, pk_column)
            query = select(model).where(pk_attr.in_(chunk_ids))
            result = await session.scalars(query)
            records = result.all()

            for record in records:
                row_data = self._record_to_dict(record, columns)
                prefix = "" if first_record else ","
                first_record = False
                yield (prefix + json.dumps(row_data, default=str)).encode("utf-8")

        yield b"]"

    @staticmethod
    def _prepare_cell_value(value: Any) -> Any:
        """Prepare a value for writing to an Excel cell.

        openpyxl handles many types natively (datetime, date, int, float, bool, str).
        This method handles special cases that need conversion.

        Args:
            value: The value to prepare.

        Returns:
            The value ready for Excel cell writing.
        """
        if value is None:
            return None
        # UUID needs to be converted to string
        if hasattr(value, "hex") and hasattr(value, "int"):
            # Duck typing check for UUID-like objects
            return str(value)
        # Decimal is handled natively by openpyxl, but we ensure it's numeric
        # datetime, date, time are handled natively by openpyxl
        # Lists and dicts should be converted to JSON strings
        if isinstance(value, (list, dict)):
            return json.dumps(value, default=str)
        # Related objects - just use string representation
        if hasattr(value, "__dict__") and not isinstance(value, (str, bytes, int, float, bool)):
            return str(value)
        return value

    async def _stream_xlsx(
        self,
        session: AsyncSession,
        model: type[Any],
        columns: list[str],
        sheet_name: str,
    ) -> AsyncGenerator[bytes, None]:
        """Stream XLSX data for all records.

        Note: openpyxl does not support true streaming, so we build the entire
        workbook in memory and yield the complete content.

        Args:
            session: The database session.
            model: The SQLAlchemy model class.
            columns: List of column names to export.
            sheet_name: Name for the worksheet (typically the model name).

        Yields:
            Complete XLSX file content as bytes.
        """
        openpyxl = _get_openpyxl()

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name[:31]  # Excel sheet names limited to 31 chars

        # Write header row
        for col_idx, column_name in enumerate(columns, start=1):
            worksheet.cell(row=1, column=col_idx, value=column_name)

        # Fetch and write all records
        row_idx = 2
        offset = 0
        while True:
            query = select(model).offset(offset).limit(CHUNK_SIZE)
            result = await session.scalars(query)
            records = result.all()

            if not records:
                break

            for record in records:
                for col_idx, column in enumerate(columns, start=1):
                    value = getattr(record, column, None)
                    worksheet.cell(row=row_idx, column=col_idx, value=self._prepare_cell_value(value))
                row_idx += 1

            offset += CHUNK_SIZE

            if len(records) < CHUNK_SIZE:
                break

        # Save to BytesIO and yield
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        yield output.getvalue()

    async def _stream_xlsx_by_ids(
        self,
        session: AsyncSession,
        model: type[Any],
        columns: list[str],
        pk_column: str,
        ids: list[Any],
        sheet_name: str,
    ) -> AsyncGenerator[bytes, None]:
        """Stream XLSX data for selected records.

        Note: openpyxl does not support true streaming, so we build the entire
        workbook in memory and yield the complete content.

        Args:
            session: The database session.
            model: The SQLAlchemy model class.
            columns: List of column names to export.
            pk_column: The primary key column name.
            ids: List of record IDs to export.
            sheet_name: Name for the worksheet (typically the model name).

        Yields:
            Complete XLSX file content as bytes.
        """
        openpyxl = _get_openpyxl()

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name[:31]  # Excel sheet names limited to 31 chars

        # Write header row
        for col_idx, column_name in enumerate(columns, start=1):
            worksheet.cell(row=1, column=col_idx, value=column_name)

        # Fetch and write selected records
        row_idx = 2
        for i in range(0, len(ids), CHUNK_SIZE):
            chunk_ids = ids[i : i + CHUNK_SIZE]
            pk_attr = getattr(model, pk_column)
            query = select(model).where(pk_attr.in_(chunk_ids))
            result = await session.scalars(query)
            records = result.all()

            for record in records:
                for col_idx, column in enumerate(columns, start=1):
                    value = getattr(record, column, None)
                    worksheet.cell(row=row_idx, column=col_idx, value=self._prepare_cell_value(value))
                row_idx += 1

        # Save to BytesIO and yield
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        yield output.getvalue()
